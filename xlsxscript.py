# import module
import openpyxl
from yt_dlp import YoutubeDL
  
  
URLS = []  
# load excel with its path
wrkbk = openpyxl.load_workbook("fn.xlsx")
  
sh = wrkbk.active
  
# iterate through excel and display data
print("READING XLSX DATA!")
for i in range(1, sh.max_row+1):
    cell_obj = sh.cell(row=i, column=1)
    print(cell_obj.value)
    URLS.append(cell_obj.value)
print("STARTING DOWNLOAD")
with YoutubeDL() as ydl:
    ydl.download(URLS)